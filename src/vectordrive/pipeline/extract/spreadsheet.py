"""Spreadsheet extraction: .xlsx via openpyxl, legacy .xls via xlrd.

A sheet has no "page" in the PDF/DOCX sense, but the existing citation
model already treats a document as a list of numbered pages — reusing it
here means every downstream stage (chunking, embedding, search, MCP
citations) needs zero changes: one sheet becomes one ExtractedPage, and a
search hit inside a spreadsheet cites "page N" the same way a PDF hit
cites a real page, with N being the sheet's 1-indexed position in the
workbook.

Each row is serialized as `col1 | col2 | col3` (blank cells rendered as
empty strings, trailing all-blank cells trimmed) so adjacent cell values
stay visually and lexically separated for both keyword and semantic
search, without inventing a table format the workbook doesn't have.
"""
from __future__ import annotations

from pathlib import Path

from vectordrive.pipeline.extract.base import ExtractedDocument, ExtractedPage


def _row_to_text(values: list) -> str:
    cells = ["" if v is None else str(v) for v in values]
    while cells and cells[-1] == "":
        cells.pop()
    return " | ".join(cells)


def _sheet_to_text(sheet_name: str, rows: list[list]) -> str:
    lines = [f"Sheet: {sheet_name}"]
    for row in rows:
        line = _row_to_text(row)
        if line:
            lines.append(line)
    return "\n".join(lines)


def _extract_xlsx(path: Path) -> list[ExtractedPage]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        pages = []
        for index, sheet_name in enumerate(wb.sheetnames, start=1):
            sheet = wb[sheet_name]
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            text = _sheet_to_text(sheet_name, rows)
            pages.append(
                ExtractedPage(page_number=index, text=text, text_source="native", needs_ocr=False)
            )
        return pages
    finally:
        wb.close()


def _extract_xls(path: Path) -> list[ExtractedPage]:
    import xlrd

    wb = xlrd.open_workbook(path)
    pages = []
    for index, sheet in enumerate(wb.sheets(), start=1):
        rows = [sheet.row_values(r) for r in range(sheet.nrows)]
        text = _sheet_to_text(sheet.name, rows)
        pages.append(
            ExtractedPage(page_number=index, text=text, text_source="native", needs_ocr=False)
        )
    return pages


class SpreadsheetExtractor:
    NAME = "spreadsheet"
    VERSION = "1"

    def extract(self, path: Path) -> ExtractedDocument:
        path = Path(path)
        if path.suffix.lower() == ".xls":
            pages = _extract_xls(path)
        else:
            pages = _extract_xlsx(path)
        return ExtractedDocument(
            extractor_name=self.NAME,
            extractor_version=self.VERSION,
            page_count=len(pages),
            pages=pages,
        )
